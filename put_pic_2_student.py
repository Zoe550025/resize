import os
import shutil
from openpyxl import load_workbook

'''讀excel'''
mywb = load_workbook('E:\\tzuwen\\tree_segmentation2\labeling\\1101-210016 資料結構與演算法(一) {mlang en} Data Structures and Algorithms (1){mlang}修課名單_1.xlsx')
sheet = mywb['1101-210016 資料結構與演算法(一) {mlang ']
#讀excel，每一欄放隨機的五個crop_list
ID_list = [[] for _ in range(64)]
#print(ID_list)
'''以 for 迴圈逐一處理每個儲存格'''
for ID in sheet.iter_cols(min_row=2, max_col=1, max_row=65, values_only=True):
    for i,cell in enumerate(ID):
        ID_list[i].append(cell)
#print(ID_list)

'''每五個addr放進每一個學生ID'''
addr_list = []
for row in sheet.iter_rows(min_row=2, min_col=3,max_col=7, max_row=65):
    for cell in row:
        addr_list.append(cell.value)
num = 0
#print(ID_list)
for index, ID in enumerate(ID_list):
    for i in range(5):
        ID_list[index].append(addr_list[num])
        num+=1
'''copyfile'''
for ID in ID_list:
    for i in range(1,6):
        shutil.copy(ID[i], os.path.join("E:\\tzuwen\\tree_segmentation2\labeling\\tree labeling2",ID[0]))
        #print(ID[i], "is copied to ", os.path.join("E:\\tzuwen\\tree_segmentation2\labeling\\tree labeling2",ID[0]))


