
from ast import literal_eval
from pathlib import Path
import cv2
import os
import shutil
import pandas as pd
import openpyxl
path_oringnal = "E:\\tzuwen\\tree_segmentation2\\resize_test\\640_20220624_2-19\\202112"#0-2\\2021\\1\\"       #TODO
excel_name = '202112_average.xlsx'
pic_list = []
pic_list1 = []
pic_name = []
dict_data = dict()

def test_folder(folderpath):
    try:
        os.makedirs(folderpath)
    # 檔案已存在的例外處理
    except FileExistsError:
        pass
        #print("檔案已存在。")

def average(xml_data,imgs):
    total = 0
    num = 0
    for xd, i in zip(xml_data, imgs):
        pic_list1 = literal_eval(xd)
        if pic_list1 != []:
            for num_box, box in enumerate(pic_list1):
                total = total + float(box[4])
                num = num + 1
    avg = total / (num)
    print( total, "/", (num), " = ", avg)
    return avg

#camera : 0-
for p_c in Path(path_oringnal).iterdir():
    if p_c.is_dir():                                        #E:\tzuwen\tree_segmentation2\resize_test\640_20220509\0\0-0
        #print(p_c)
        for p_y in Path(p_c).iterdir():                     #E:\tzuwen\tree_segmentation2\resize_test\640_20220509\0\0-0\2021
            #print(p_y)
            if p_y.is_dir():
                #print(p_y)
                for p_m in Path(p_y).iterdir():             #E:\tzuwen\tree_segmentation2\resize_test\640_20220509\0\0-0\2021\1
                    if p_m.is_dir():
                        #print(p_m)
                        tmp2 = str(p_m).split("\\")  # ['E:', 'tzuwen', 'tree_segmentation2', 'resize_test', '640_20220509', '0', '0-9', '2022', '5']

                    ##pic :
                        for p in Path(p_m).iterdir():       #E:\tzuwen\tree_segmentation2\resize_test\640_20220509\0\0-0\2021\1\2021-01-27_1552.png
                            if p.is_file():
                                tmp = str(p).split("\\")    #['E:', 'tzuwen', 'tree_segmentation2', 'resize_test', '640_20220509', '0', '0-81', '2022', '4', '2022-04-05_1042.png']
                                #print(tmp)
                                pic_list.append(str(p))
                                pic_name.append(tmp[-1])
                        try:
                            with open(os.path.join(p_m,"results\\coordinates.json"),"r") as f:
                                data = f.readlines()
                        except:
                            pass
                        month_average = average(data, pic_name)
                        print(month_average)
                        if tmp2[6] in dict_data:
                            dict_data[tmp2[6]] += [[tmp2[7],tmp2[8],month_average]]
                        else:
                            dict_data[tmp2[6]] = [[tmp2[7],tmp2[8],month_average]]

                        for d,p in zip(data,pic_name):
                           pic_list=literal_eval(d)
                           if pic_list != []:
                               for num_box,box in enumerate(pic_list):
                                   #print(num_box,p)
                                   if int(box[4])> month_average:          #TODO:要crop的average準確率值
                                       ymin = int(box[0])
                                       ymax = int(box[1])
                                       xmin = int(box[2])
                                       xmax = int(box[3])
                                       img_path = os.path.join(str(p_m),str(p))
                                       img = cv2.imread(img_path)
                                       print(img_path)
                                       crop_img = img[ymin:ymax,xmin:xmax]
                                       shutil.rmtree(os.path.join(p_m,"crop/"), ignore_errors=True)
                                       test_folder(os.path.join(p_m,"crop/"))
                                       tmp_str = "crop/"+ p.replace(".png","") + "_" + str(num_box)+".png"
                                       cv2.imwrite(os.path.join(p_m,tmp_str),crop_img)

#build an excel
key_table = dict()
def colA_and_row2_init():
    # use dict_data.keys to write column A, start from A3
    for i, key in enumerate(dict_data.keys(), start=3):
        A_tmp = 'A' + str(i)
        actSheet[A_tmp] = key  # A3 -> 0-0
        key_table[key] = str(i)  #{'0-0': '3', '0-1': '4', '0-10': '5', '0-11': '6',....}


    # write month at row2
    months = [1,2,3,4,5,6,7,8,9,10,11,12]
    for j, month in enumerate(months,start=1):
        row2_tmp = num2ord(j)+str(2)
        actSheet[row2_tmp] = month

def num2ord(month):  #1 to colB
    return chr(int(month)+65)

wb = openpyxl.Workbook()

# print(actSheet.max_row)
# print(actSheet.max_column)

for key in dict_data.keys():
    for val in dict_data[key]:      #['2021', '11', 11.56770410843646]
        if val[0] not in wb.sheetnames:
            wb.create_sheet(val[0])
            actSheet = wb[val[0]]
            actSheet['B1'] = val[0]
            colA_and_row2_init()
        else:
            actSheet = wb[val[0]]
            actSheet['B1'] = val[0]

        col_val = num2ord(int(val[1]))
        row_val = key_table.get(key)
        excel_val = str(col_val) + str(row_val)
        actSheet[excel_val] = val[2]

wb.save(excel_name)